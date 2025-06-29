from flask import Blueprint, request, jsonify
import openpyxl
import os
from datetime import datetime
from src.models.user import db
from src.models.venda import Venda
from src.models.vendedor_comissao import VendedorComissao
from src.models.vendedor import Vendedor

importacao_bp = Blueprint('importacao', __name__)

def validar_cpf(cpf):
    """Validação básica de CPF (formato)"""
    if not cpf:
        return False
    
    # Remover caracteres especiais
    cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
    
    # Verificar se tem 11 dígitos
    if len(cpf_limpo) != 11:
        return False
    
    # Verificar se não são todos os dígitos iguais
    if cpf_limpo == cpf_limpo[0] * 11:
        return False
    
    return True

def formatar_cpf(cpf):
    """Formatar CPF para o padrão XXX.XXX.XXX-XX"""
    if not cpf:
        return None
    
    cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
    if len(cpf_limpo) == 11:
        return f"{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:]}"
    return cpf

@importacao_bp.route('/importacao/template', methods=['GET'])
def baixar_template():
    """Gera e retorna um template Excel para importação de vendas"""
    try:
        # Buscar vendedores e suas configurações de comissão para o exemplo
        vendedores_comissoes = VendedorComissao.query.join(Vendedor).all()
        
        # Criar workbook
        wb = openpyxl.Workbook()
        
        # Aba principal com dados de exemplo
        ws_vendas = wb.active
        ws_vendas.title = "Vendas"
        
        # Cabeçalhos
        headers = ['cpf_cliente', 'nome_cliente', 'data_venda', 'valor_venda', 'vendedor', 'tabela_comissao', 'usuario_cadastro']
        for col, header in enumerate(headers, 1):
            ws_vendas.cell(row=1, column=col, value=header)
        
        # Dados de exemplo
        if vendedores_comissoes:
            for i, vc in enumerate(vendedores_comissoes[:3], 2):  # Começar na linha 2
                ws_vendas.cell(row=i, column=1, value=f'123.456.789-{i-2:02d}')
                ws_vendas.cell(row=i, column=2, value=f'Cliente Exemplo {i-1}')
                ws_vendas.cell(row=i, column=3, value='2025-06-29')
                ws_vendas.cell(row=i, column=4, value=1000.00 * (i-1))
                ws_vendas.cell(row=i, column=5, value=vc.vendedor.nome_vendedor)
                ws_vendas.cell(row=i, column=6, value=vc.nome_tabela)
                ws_vendas.cell(row=i, column=7, value='admin')
        else:
            # Exemplo genérico
            ws_vendas.cell(row=2, column=1, value='123.456.789-00')
            ws_vendas.cell(row=2, column=2, value='Cliente Exemplo')
            ws_vendas.cell(row=2, column=3, value='2025-06-29')
            ws_vendas.cell(row=2, column=4, value=1000.00)
            ws_vendas.cell(row=2, column=5, value='Nome do Vendedor')
            ws_vendas.cell(row=2, column=6, value='Nome da Tabela')
            ws_vendas.cell(row=2, column=7, value='admin')
        
        # Aba de instruções
        ws_instrucoes = wb.create_sheet("Instruções")
        instrucoes_data = [
            ['Campo', 'Descrição', 'Obrigatório', 'Exemplo'],
            ['cpf_cliente', 'CPF do cliente (formato: XXX.XXX.XXX-XX ou apenas números)', 'Sim', '123.456.789-00'],
            ['nome_cliente', 'Nome completo do cliente', 'Sim', 'João Silva'],
            ['data_venda', 'Data da venda (formato: AAAA-MM-DD)', 'Sim', '2025-06-29'],
            ['valor_venda', 'Valor da venda (formato: 1000.50)', 'Sim', '1500.00'],
            ['vendedor', 'Nome exato do vendedor (deve existir no sistema)', 'Sim', 'Ana Silva'],
            ['tabela_comissao', 'Nome da tabela de comissão (deve existir para o vendedor)', 'Sim', 'Gold'],
            ['usuario_cadastro', 'Usuário que está cadastrando (opcional)', 'Não', 'admin']
        ]
        
        for row, data in enumerate(instrucoes_data, 1):
            for col, value in enumerate(data, 1):
                ws_instrucoes.cell(row=row, column=col, value=value)
        
        # Aba com vendedores disponíveis
        if vendedores_comissoes:
            ws_vendedores = wb.create_sheet("Vendedores_Disponíveis")
            ws_vendedores.cell(row=1, column=1, value='Vendedor')
            ws_vendedores.cell(row=1, column=2, value='Tabela_Comissao')
            ws_vendedores.cell(row=1, column=3, value='Porcentagem')
            
            for row, vc in enumerate(vendedores_comissoes, 2):
                ws_vendedores.cell(row=row, column=1, value=vc.vendedor.nome_vendedor)
                ws_vendedores.cell(row=row, column=2, value=vc.nome_tabela)
                ws_vendedores.cell(row=row, column=3, value=f"{vc.porcentagem_comissao}%")
        
        # Salvar arquivo
        temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        arquivo_template = os.path.join(temp_dir, 'template_vendas.xlsx')
        wb.save(arquivo_template)
        
        return jsonify({
            'mensagem': 'Template gerado com sucesso',
            'arquivo': arquivo_template,
            'download_url': f'/api/importacao/template/download'
        }), 200
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@importacao_bp.route('/importacao/template/download', methods=['GET'])
def download_template():
    """Download do arquivo template"""
    try:
        from flask import send_file
        
        arquivo_template = os.path.join(os.path.dirname(__file__), '..', 'temp', 'template_vendas.xlsx')
        
        if not os.path.exists(arquivo_template):
            return jsonify({'erro': 'Template não encontrado. Gere o template primeiro.'}), 404
        
        return send_file(
            arquivo_template,
            as_attachment=True,
            download_name='template_vendas.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@importacao_bp.route('/importacao/vendas', methods=['POST'])
def importar_vendas():
    """Importa vendas de um arquivo Excel"""
    try:
        # Verificar se foi enviado um arquivo
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
        
        # Verificar extensão do arquivo
        if not arquivo.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        # Salvar arquivo temporariamente
        temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        arquivo_path = os.path.join(temp_dir, f'upload_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{arquivo.filename}')
        arquivo.save(arquivo_path)
        
        try:
            # Ler arquivo Excel
            wb = openpyxl.load_workbook(arquivo_path)
            ws = wb.active
            
            # Ler cabeçalhos
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())
            
            # Verificar colunas obrigatórias
            colunas_obrigatorias = ['cpf_cliente', 'nome_cliente', 'data_venda', 'valor_venda', 'vendedor', 'tabela_comissao']
            colunas_faltantes = [col for col in colunas_obrigatorias if col not in headers]
            
            if colunas_faltantes:
                return jsonify({
                    'erro': f'Colunas obrigatórias faltantes: {", ".join(colunas_faltantes)}'
                }), 400
            
            # Mapear índices das colunas
            col_indices = {header: headers.index(header) + 1 for header in headers}
            
            # Processar cada linha
            vendas_criadas = []
            erros = []
            
            for row_num in range(2, ws.max_row + 1):  # Começar da linha 2 (pular cabeçalho)
                try:
                    # Ler dados da linha
                    row_data = {}
                    for header in headers:
                        cell_value = ws.cell(row=row_num, column=col_indices[header]).value
                        row_data[header] = cell_value
                    
                    # Validar dados obrigatórios
                    if not row_data.get('cpf_cliente') or not row_data.get('nome_cliente') or not row_data.get('valor_venda'):
                        erros.append(f'Linha {row_num}: Dados obrigatórios faltantes')
                        continue
                    
                    # Validar e formatar CPF
                    cpf = str(row_data['cpf_cliente']).strip()
                    if not validar_cpf(cpf):
                        erros.append(f'Linha {row_num}: CPF inválido ({cpf})')
                        continue
                    
                    cpf_formatado = formatar_cpf(cpf)
                    
                    # Validar valor da venda
                    try:
                        valor_venda = float(row_data['valor_venda'])
                        if valor_venda <= 0:
                            erros.append(f'Linha {row_num}: Valor da venda deve ser maior que zero')
                            continue
                    except (ValueError, TypeError):
                        erros.append(f'Linha {row_num}: Valor da venda inválido')
                        continue
                    
                    # Buscar vendedor e configuração de comissão
                    nome_vendedor = str(row_data['vendedor']).strip()
                    nome_tabela = str(row_data['tabela_comissao']).strip()
                    
                    vendedor_comissao = db.session.query(VendedorComissao).join(Vendedor).filter(
                        Vendedor.nome_vendedor == nome_vendedor,
                        VendedorComissao.nome_tabela == nome_tabela
                    ).first()
                    
                    if not vendedor_comissao:
                        erros.append(f'Linha {row_num}: Vendedor "{nome_vendedor}" com tabela "{nome_tabela}" não encontrado')
                        continue
                    
                    # Processar data da venda
                    data_venda = None
                    if row_data.get('data_venda'):
                        try:
                            if isinstance(row_data['data_venda'], str):
                                data_venda = datetime.strptime(row_data['data_venda'], '%Y-%m-%d').date()
                            else:
                                data_venda = row_data['data_venda'].date() if hasattr(row_data['data_venda'], 'date') else row_data['data_venda']
                        except (ValueError, AttributeError):
                            erros.append(f'Linha {row_num}: Data inválida (use formato AAAA-MM-DD)')
                            continue
                    
                    # Criar venda
                    venda = Venda(
                        cpf_cliente=cpf_formatado,
                        nome_cliente=str(row_data['nome_cliente']).strip(),
                        valor_venda=valor_venda,
                        id_vendedor_comissao=vendedor_comissao.id,
                        usuario_cadastro=str(row_data.get('usuario_cadastro', 'Importação Excel')).strip(),
                        valor_comissao=0.0  # Será calculado automaticamente
                    )
                    
                    if data_venda:
                        venda.data_venda = data_venda
                    
                    # Adicionar ao banco
                    db.session.add(venda)
                    db.session.flush()  # Para ter acesso ao relacionamento
                    
                    # Calcular comissão
                    venda.calcular_comissao()
                    
                    vendas_criadas.append({
                        'linha': row_num,
                        'cliente': venda.nome_cliente,
                        'valor': venda.valor_venda,
                        'comissao': venda.valor_comissao,
                        'vendedor': vendedor_comissao.vendedor.nome_vendedor,
                        'tabela': vendedor_comissao.nome_tabela
                    })
                    
                except Exception as e:
                    erros.append(f'Linha {row_num}: Erro ao processar - {str(e)}')
                    continue
            
            # Commit apenas se houver vendas criadas
            if vendas_criadas:
                db.session.commit()
            else:
                db.session.rollback()
            
            # Limpar arquivo temporário
            if os.path.exists(arquivo_path):
                os.remove(arquivo_path)
            
            return jsonify({
                'mensagem': f'Importação concluída',
                'vendas_criadas': len(vendas_criadas),
                'erros_encontrados': len(erros),
                'detalhes': {
                    'vendas': vendas_criadas,
                    'erros': erros
                }
            }), 200
            
        except Exception as e:
            db.session.rollback()
            # Limpar arquivo temporário
            if os.path.exists(arquivo_path):
                os.remove(arquivo_path)
            raise e
            
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

